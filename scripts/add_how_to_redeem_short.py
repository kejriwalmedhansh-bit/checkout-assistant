"""
Generate `how_to_redeem_short` for every brand in data/gyftr_master.json using
Claude Haiku 4.5, then mirror the result into db/gyftr_master.json and add a
matching column to gyftr_full_database.xlsx.

Resumable: brands that already have `how_to_redeem_short` set (including a
value of null) are skipped. Progress is saved to disk after every brand.

Usage:
    python3.11 scripts/add_how_to_redeem_short.py             # process all brands
    python3.11 scripts/add_how_to_redeem_short.py --limit 10  # process only 10 (test run)
"""

import argparse
import json
import time
from pathlib import Path
from typing import Optional

import anthropic
from dotenv import load_dotenv
from openpyxl import load_workbook
from pydantic import BaseModel

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_PATH = REPO_ROOT / "data" / "gyftr_master.json"
DB_PATH = REPO_ROOT / "db" / "gyftr_master.json"
XLSX_PATH = REPO_ROOT / "gyftr_full_database.xlsx"

MODEL = "claude-haiku-4-5"
DELAY_SECONDS = 0.4
XLSX_HEADER = "How to Redeem (Short)"

SYSTEM_PROMPT = """You write extremely short redemption instructions for an Indian e-commerce checkout assistant app. Given a brand's existing (long, unstructured) redemption steps and its redemption type, produce ONE short instruction following these exact rules:

- Max 2 lines, plain text, no bullet points, no numbering.
- If redemption_type is "Online" or "Both": condense ONLY the online redemption steps.
- If redemption_type is "Offline": condense ONLY the offline (in-store) steps.
- Focus ONLY on: where to find the gift card/voucher entry field, and what to enter there.
- Do NOT include URLs, "visit website", or login/register instructions — assume the user is already on the site/app.
- Plain, simple Indian English. No corporate speak, no marketing language.
- Online pattern to follow: "[Where to find it] -> enter voucher code + PIN" (adapt wording to what the steps actually say -- some brands use a mobile number instead of a PIN; keep the arrow pattern).
- Offline pattern to follow: "Show voucher code to cashier before billing" (adapt if the brand's steps describe a different offline flow, e.g. sharing a mobile number instead of a code).

If the steps genuinely don't contain enough information to produce a useful instruction, return null instead of guessing."""


class RedeemShort(BaseModel):
    how_to_redeem_short: Optional[str] = None


def needs_generation(brand: dict) -> bool:
    return "how_to_redeem_short" not in brand


def has_usable_steps(brand: dict) -> bool:
    steps = brand.get("how_to_redeem_steps")
    return bool(steps)


def build_user_message(brand: dict) -> str:
    steps = brand.get("how_to_redeem_steps") or []
    steps_text = "\n".join(f"{i + 1}. {s}" for i, s in enumerate(steps))
    return (
        f"Brand: {brand.get('brand_name')}\n"
        f"Redemption type: {brand.get('redemption_type')}\n"
        f"Existing redemption steps:\n{steps_text}"
    )


def generate_short_instruction(client: anthropic.Anthropic, brand: dict) -> Optional[str]:
    response = client.messages.parse(
        model=MODEL,
        max_tokens=300,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": build_user_message(brand)}],
        output_format=RedeemShort,
    )
    return response.parsed_output.how_to_redeem_short


def save_json(path: Path, data: dict) -> None:
    with open(path, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def sync_db_copy(data: dict) -> None:
    if not DB_PATH.exists():
        print(f"  (skipping db/ mirror -- {DB_PATH} not found)")
        return
    with open(DB_PATH) as f:
        db_data = json.load(f)

    updated = 0
    for slug, brand in data.items():
        if "how_to_redeem_short" in brand and slug in db_data:
            db_data[slug]["how_to_redeem_short"] = brand["how_to_redeem_short"]
            updated += 1

    save_json(DB_PATH, db_data)
    print(f"Mirrored how_to_redeem_short into {DB_PATH} for {updated} brands.")


def sync_xlsx(data: dict) -> None:
    if not XLSX_PATH.exists():
        print(f"  (skipping xlsx update -- {XLSX_PATH} not found)")
        return
    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    header_row = ws[1]
    slug_col = None
    short_col = None
    for cell in header_row:
        if cell.value == "Slug":
            slug_col = cell.column
        elif cell.value == XLSX_HEADER:
            short_col = cell.column

    if slug_col is None:
        print(f"  (skipping xlsx update -- no 'Slug' column found in {XLSX_PATH})")
        return

    if short_col is None:
        short_col = ws.max_column + 1
        ws.cell(row=1, column=short_col, value=XLSX_HEADER)

    updated = 0
    for row in ws.iter_rows(min_row=2):
        slug_cell = row[slug_col - 1]
        slug = slug_cell.value
        brand = data.get(slug)
        if brand is not None and "how_to_redeem_short" in brand:
            ws.cell(row=slug_cell.row, column=short_col, value=brand["how_to_redeem_short"] or "")
            updated += 1

    wb.save(XLSX_PATH)
    print(f"Updated '{XLSX_HEADER}' column in {XLSX_PATH} for {updated} brands.")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=None, help="Only process this many brands (for testing)")
    args = parser.parse_args()

    load_dotenv()
    client = anthropic.Anthropic()

    with open(DATA_PATH) as f:
        data = json.load(f)

    pending = [(slug, brand) for slug, brand in data.items() if needs_generation(brand)]
    total_pending = len(pending)
    if args.limit is not None:
        pending = pending[: args.limit]

    print(f"{total_pending} brands need how_to_redeem_short. Processing {len(pending)} now.\n")

    generated = 0
    skipped_no_steps = 0
    errors = 0

    for i, (slug, brand) in enumerate(pending, 1):
        if not has_usable_steps(brand):
            brand["how_to_redeem_short"] = None
            skipped_no_steps += 1
            print(f"[{i}/{len(pending)}] {slug}: SKIPPED (no steps) -> null")
        else:
            try:
                short_text = generate_short_instruction(client, brand)
                brand["how_to_redeem_short"] = short_text
                generated += 1
                print(f"[{i}/{len(pending)}] {slug}: {short_text}")
            except Exception as e:
                errors += 1
                print(f"[{i}/{len(pending)}] {slug}: ERROR ({e}) -- will retry next run")
                time.sleep(DELAY_SECONDS)
                continue

        save_json(DATA_PATH, data)
        time.sleep(DELAY_SECONDS)

    print(
        f"\nDone. Generated: {generated}, skipped (no steps): {skipped_no_steps}, "
        f"errors (will retry next run): {errors}."
    )

    sync_db_copy(data)
    sync_xlsx(data)


if __name__ == "__main__":
    main()
