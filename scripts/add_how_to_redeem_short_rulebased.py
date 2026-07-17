"""
Generate `how_to_redeem_short` for every brand in db/gyftr_master.json using
pure Python string/regex matching -- no AI calls, no network calls.

Rules:
1. redemption_type == "Offline" -> fixed string: "Show voucher code to cashier before billing"
2. redemption_type in ("Online", "Both") -> scan how_to_redeem_steps for the
   sentence describing entering a gift card / voucher / e-pay / wallet code,
   and format as "[Where] -> enter voucher code + PIN"
3. how_to_redeem_steps missing/empty, or containing "NEEDS REVIEW" -> null
   (this takes priority over rule 1 -- an Offline brand with no step data
   also gets null, not the fixed string)

Cleaning: strip URLs, strip login/register/sign-up instructions, strip
"visit website" instructions, no bullet points, plain simple English.

Usage:
    python3.11 scripts/add_how_to_redeem_short_rulebased.py --preview   # print first 15, write nothing
    python3.11 scripts/add_how_to_redeem_short_rulebased.py             # write for real
"""

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
DB_PATH = REPO_ROOT / "db" / "gyftr_master.json"
DATA_PATH = REPO_ROOT / "data" / "gyftr_master.json"
XLSX_PATH = REPO_ROOT / "gyftr_full_database.xlsx"

OFFLINE_FIXED = "Show voucher code to cashier before billing"
XLSX_HEADER = "How to Redeem (Short)"

URL_RE = re.compile(r"(https?://\S+|www\.\S+|\b[a-z0-9][a-z0-9.-]*\.(?:com|in|co|io|net|org)\b\S*)", re.I)
LOGIN_RE = re.compile(r"\b(log[\s-]?in|sign[\s-]?up|sign[\s-]?in|register|create an? account)\b", re.I)
VISIT_RE = re.compile(r"\b(visit|download|open)\s+(the\s+)?(website|app|application|store|page)\b", re.I)
# In-person/cashier flow descriptions -- these describe the *offline* half of a
# "Both" brand's steps and must never be picked as the online "[Where]" candidate.
OFFLINE_INDICATOR_RE = re.compile(
    r"\b(cashier|before\s+billing|at\s+the\s+time\s+of\s+billing|in-?store|at\s+store|outlet\s+locator|"
    r"nearest\s+outlet|offline\s+redemption|for\s+offline|cash\s+counter)\b",
    re.I,
)
BULLET_LEAD_RE = re.compile(r"^[\-\*•]\s*")
BULLET_MID_RE = re.compile(r"(?<!\w)-\s*(?=[A-Z])")

VOUCHER_KEYWORDS = re.compile(
    r"\b(gift\s*card|gift\s*voucher|voucher|coupon|e-?pay|wallet|gv|gc|promo\s*code|discount\s*code|code)\b",
    re.I,
)
ACTION_KEYWORDS = re.compile(
    r"\b(enter|add|apply|select|choose|click|tap|redeem|input)\b", re.I
)
ENTER_RE = re.compile(r"\benter\b", re.I)


def clean_text(s: str) -> str:
    if not s:
        return ""
    s = URL_RE.sub("", s)
    s = BULLET_LEAD_RE.sub("", s.strip())
    s = re.sub(r"\s+", " ", s).strip(" .,:;-&")
    return s


def split_sentences(step: str) -> list[str]:
    step = BULLET_MID_RE.sub(". ", step)
    # Split on sentence-ending punctuation followed by a capital letter,
    # tolerating source-data typos with no space after the period.
    parts = re.split(r"(?<=[.!])\s*(?=[A-Z])", step)
    return [p.strip() for p in parts if p.strip()]


def strip_disallowed(sentence: str) -> str:
    """Truncate at the first login/visit/URL mention rather than discarding
    the whole sentence -- some steps legitimately describe the voucher entry
    action before or after an unrelated login/visit clause."""
    cut_points = [m.start() for m in (LOGIN_RE.search(sentence), VISIT_RE.search(sentence), URL_RE.search(sentence)) if m]
    if cut_points:
        sentence = sentence[: min(cut_points)]
    return sentence.strip(" ,.;:-")


FILLER_WHERE = {"you can", "you will", "please", "kindly", "customer to", "customer will", "user can"}


def extract_where(sentence: str) -> str:
    sentence = clean_text(sentence)
    match = ENTER_RE.search(sentence) or ACTION_KEYWORDS.search(sentence)
    where = sentence[: match.start()] if match and match.start() > 3 else ""
    where = clean_text(where)
    where = re.sub(r"\b(and|to|the|or)$", "", where, flags=re.I).strip(" ,.;:-&")
    if len(where) < 3 or where.lower() in FILLER_WHERE:
        return "At checkout"
    return where[0].upper() + where[1:]


def find_candidate(steps: list[str]) -> str | None:
    """Collect every sentence that mentions a voucher word alongside an action
    word, ranked "enter"+voucher first (the strongest signal for the
    template) then any other action+voucher. Within each rank, prefer a
    sentence that actually yields a real [Where] location over one that
    falls back to the generic "At checkout" -- only settle for a generic
    result if nothing better exists anywhere in the steps."""
    ranked: list[tuple[str, int]] = []
    for step in steps:
        for sent in split_sentences(step):
            if OFFLINE_INDICATOR_RE.search(sent):
                continue
            sent = strip_disallowed(sent)
            if not sent or not VOUCHER_KEYWORDS.search(sent):
                continue
            if ENTER_RE.search(sent):
                ranked.append((sent, 0))
            elif ACTION_KEYWORDS.search(sent):
                ranked.append((sent, 1))
    if not ranked:
        return None
    ranked.sort(key=lambda pair: pair[1])
    for sent, _ in ranked:
        if extract_where(sent) != "At checkout":
            return sent
    return ranked[0][0]


def has_needs_review(steps) -> bool:
    return any("NEEDS REVIEW" in str(s).upper() for s in (steps or []))


def rule_based_short(brand: dict) -> tuple[str | None, str]:
    """Returns (value, category) where category explains how it was derived."""
    steps = brand.get("how_to_redeem_steps")
    rtype = brand.get("redemption_type")

    if not steps or has_needs_review(steps):
        return None, "null_missing_or_needs_review"

    if rtype == "Offline":
        return OFFLINE_FIXED, "offline_fixed"

    if rtype in ("Online", "Both"):
        candidate = find_candidate(steps)
        if candidate is None:
            return f"At checkout → enter voucher code + PIN", "online_fallback"
        where = extract_where(candidate)
        return f"{where} → enter voucher code + PIN", "online_extracted"

    return None, "null_missing_or_needs_review"


def process_all(data: dict) -> dict:
    results = {}
    for slug, brand in data.items():
        value, category = rule_based_short(brand)
        results[slug] = (value, category)
    return results


def sync_data_copy(db_data: dict) -> None:
    if not DATA_PATH.exists():
        print(f"  (skipping data/ mirror -- {DATA_PATH} not found)")
        return
    with open(DATA_PATH) as f:
        data_data = json.load(f)

    updated = 0
    for slug, brand in db_data.items():
        if slug in data_data:
            data_data[slug]["how_to_redeem_short"] = brand.get("how_to_redeem_short")
            updated += 1

    with open(DATA_PATH, "w") as f:
        json.dump(data_data, f, indent=2, ensure_ascii=False)
    print(f"Mirrored how_to_redeem_short into {DATA_PATH} for {updated} brands.")


def sync_xlsx(db_data: dict) -> None:
    if not XLSX_PATH.exists():
        print(f"  (skipping xlsx update -- {XLSX_PATH} not found)")
        return
    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    header_row = ws[1]
    slug_col = None
    short_col = None
    for cell in header_row:
        if cell.value == "Slug":
            slug_col = cell.column
        elif cell.value == XLSX_HEADER:
            short_col = cell.column

    if slug_col is None:
        print(f"  (skipping xlsx update -- no 'Slug' column found in {XLSX_PATH})")
        return

    if short_col is None:
        short_col = ws.max_column + 1
        ws.cell(row=1, column=short_col, value=XLSX_HEADER)

    updated = 0
    for row in ws.iter_rows(min_row=2):
        slug_cell = row[slug_col - 1]
        brand = db_data.get(slug_cell.value)
        if brand is not None:
            ws.cell(row=slug_cell.row, column=short_col, value=brand.get("how_to_redeem_short") or "")
            updated += 1

    wb.save(XLSX_PATH)
    print(f"Updated '{XLSX_HEADER}' column in {XLSX_PATH} for {updated} brands.")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--preview", action="store_true", help="Print first 15 results only, write nothing")
    args = parser.parse_args()

    with open(DB_PATH) as f:
        db_data = json.load(f)

    results = process_all(db_data)

    counts = {}
    for _, category in results.values():
        counts[category] = counts.get(category, 0) + 1

    items = list(db_data.items())
    preview_n = 15 if args.preview else len(items)
    for slug, brand in items[:preview_n]:
        value, category = results[slug]
        print(f"[{category}] {slug} ({brand.get('redemption_type')!r}): {value}")

    print()
    print("Counts:", counts)

    if args.preview:
        print("\nPreview mode -- nothing was written.")
        return

    for slug, brand in db_data.items():
        brand["how_to_redeem_short"] = results[slug][0]

    with open(DB_PATH, "w") as f:
        json.dump(db_data, f, indent=2, ensure_ascii=False)
    print(f"\nWrote how_to_redeem_short into {DB_PATH} for {len(db_data)} brands.")

    sync_data_copy(db_data)
    sync_xlsx(db_data)


if __name__ == "__main__":
    main()
