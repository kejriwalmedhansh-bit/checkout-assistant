import json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ILLEGAL = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')
def clean(text):
    if text is None:
        return ""
    return ILLEGAL.sub('', str(text)).strip()

with open("db/gyftr_master.json") as f:
    master = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = "All Brands"

headers = [
    "Brand Name", "Slug", "Redemption Type",
    "Discount % (Credit/Debit Card)", "Discount % (Net Banking)", "Discount % (UPI)",
    "Best Payment Method", "Best Discount %",
    "Min Denomination", "Max Denomination",
    "Stack Limit (max per bill)", "Stack Limit Confidence",
    "Can Club With Offers?", "One-Time Use?",
    "Redemption Restrictions", "How to Redeem (steps)",
    "Status", "Notes", "Last Scraped",
]
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="2F5233")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for slug, m in sorted(master.items(), key=lambda kv: kv[1].get("brand_name") or kv[0]):
    denoms = m.get("denominations") or []
    min_d = min(denoms) if denoms else ""
    max_d = max(denoms) if denoms else ""

    restrictions = m.get("redemption_restrictions") or []
    restrictions_text = "; ".join(restrictions) if restrictions else ""

    steps = m.get("how_to_redeem_steps")
    if steps is None:
        steps_text = "NEEDS REVIEW - not parsed"
    elif not steps:
        steps_text = ""
    else:
        steps_text = "\n".join(f"{i+1}. {s}" for i, s in enumerate(steps))

    can_club = m.get("can_club_with_offers")
    can_club_text = "Yes" if can_club is True else "No" if can_club is False else "Unknown"

    stack_limit = m.get("stack_limit")
    if stack_limit is not None:
        stack_limit_text = stack_limit
    elif m.get("stack_limit_confidence") == "unlimited_stated":
        stack_limit_text = "Unlimited (stated)"
    else:
        stack_limit_text = "Unknown"

    row = [
        clean(m.get("brand_name")), slug, clean(m.get("redemption_type")),
        m.get("discounts", {}).get("Credit/Debit Card"),
        m.get("discounts", {}).get("Net Banking"),
        m.get("discounts", {}).get("UPI"),
        clean(m.get("best_payment_method")), m.get("best_discount_pct"),
        min_d, max_d,
        stack_limit_text, clean(m.get("stack_limit_confidence")),
        can_club_text, "Yes" if m.get("one_time_use") else "No",
        clean(restrictions_text), clean(steps_text),
        clean(m.get("status")), clean(m.get("notes")), clean(m.get("last_scraped")),
    ]
    ws.append(row)

widths = [26, 26, 12, 16, 14, 10, 16, 12, 12, 12, 16, 16, 14, 12, 45, 45, 14, 20, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

ws.freeze_panes = "C2"
ws.auto_filter.ref = ws.dimensions

wb.save("gyftr_full_database.xlsx")
print("Saved gyftr_full_database.xlsx from db/gyftr_master.json")
print(f"Total brands: {len(master)}")
