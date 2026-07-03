import json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ILLEGAL = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def clean(text):
    if not text:
        return ""
    return ILLEGAL.sub('', str(text)).strip()

with open("db/gyftr_vouchers.json") as f:
    brands = json.load(f)

with open("db/gyftr_full_scrape_progress.json") as f:
    scraped = json.load(f)

REDEMPTION_MAP = {"ON": "Online", "OFF": "Offline", "B": "Both"}
STACKING_KEYWORDS = ["per bill", "per day", "only 1", "only 2", "only 3", "only one",
                     "cannot be used multiple", "cannot be clubbed", "one-time use",
                     "one time use", "single use"]

def get_discount(pgdis_list, target_names):
    if not pgdis_list:
        return None
    for pg in pgdis_list:
        name = (pg.get("pg_name") or "").strip().lower()
        if any(t.lower() in name for t in target_names):
            return pg.get("brand_pg_discount", None)
    return None

def best_payment_method(pgdis_list):
    if not pgdis_list:
        return "", ""
    best_val = max((pg.get("brand_pg_discount", 0) or 0) for pg in pgdis_list)
    tied = [pg for pg in pgdis_list if (pg.get("brand_pg_discount", 0) or 0) == best_val]
    if len(tied) > 1:
        upi_pick = next((pg for pg in tied if "upi" in (pg.get("pg_name") or "").lower()), None)
        chosen = upi_pick or tied[0]
    else:
        chosen = tied[0]
    return chosen.get("pg_name", ""), best_val

def stacking_signal(text_blocks):
    combined = " ".join(t.lower() for t in text_blocks if t)
    return "Yes" if any(kw in combined for kw in STACKING_KEYWORDS) else "No"

wb = Workbook()
ws = wb.active
ws.title = "All Brands"

headers = [
    "Brand Name", "Slug", "Redemption Type",
    "Discount % (Credit/Debit Card)", "Discount % (Net Banking)",
    "Discount % (UPI)", "Discount % (Paytm UPI)", "Discount % (Amazon Pay)",
    "Best Payment Method", "Best Discount %",
    "Min Denomination", "Max Denomination",
    "Stacking Limit Mentioned?",
    "How to Redeem", "T&Cs Summary (Important Instructions)",
    "Full Terms & Conditions", "Watch Video",
    "Status", "Notes", "Last Scraped"
]
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color="2F5233")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for brand in sorted(brands, key=lambda b: b.get("brand_name", b.get("slug", ""))):
    slug = brand.get("slug", "")
    name = brand.get("brand_name", slug)
    redemption_type = REDEMPTION_MAP.get(brand.get("redemption_type", ""), brand.get("redemption_type", ""))

    pgdis = brand.get("pgdis", [])
    cc_dc = get_discount(pgdis, ["credit card", "debit card"])
    net_banking = get_discount(pgdis, ["net banking", "netbanking"])
    upi = get_discount(pgdis, ["upi"])
    paytm_upi = get_discount(pgdis, ["paytm upi"])
    amazon_pay = get_discount(pgdis, ["amazon pay"])
    best_method, best_disc = best_payment_method(pgdis)

    products = brand.get("products", [])
    denom_values = [p.get("mrp") for p in products if p.get("mrp")]
    min_denom = min(denom_values) if denom_values else ""
    max_denom = max(denom_values) if denom_values else ""

    yt_id = brand.get("youtube_video", "")
    video_link = f"https://www.youtube.com/watch?v={yt_id}" if yt_id else ""

    scrape_data = scraped.get(slug, {})
    if scrape_data.get("delisted"):
        status = "Delisted"
        notes = scrape_data.get("reason", "")
    elif scrape_data.get("skipped"):
        status = "Skipped (out of scope)"
        notes = scrape_data.get("reason", "")
    elif scrape_data.get("error"):
        status = "Scrape Failed"
        notes = scrape_data.get("error", "")
    elif scrape_data:
        status = "Scraped"
        notes = ""
    else:
        status = "Not yet scraped"
        notes = ""

    how_to_redeem = clean(scrape_data.get("How to use", ""))
    tc_summary = clean(scrape_data.get("Important Instructions", "")) or clean(brand.get("important_instruction", ""))
    full_tc = clean(scrape_data.get("Terms & Conditions", ""))
    last_scraped = scrape_data.get("scraped_at", "")

    stacking = stacking_signal([tc_summary, full_tc])

    row = [
        clean(name), slug, redemption_type,
        cc_dc, net_banking, upi, paytm_upi, amazon_pay,
        best_method, best_disc,
        min_denom, max_denom,
        stacking,
        how_to_redeem, tc_summary, full_tc, video_link,
        status, clean(notes), last_scraped,
    ]
    ws.append(row)

widths = [26, 26, 12, 14, 12, 8, 12, 12, 16, 12, 12, 12, 14, 42, 42, 50, 30, 16, 35, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

ws.freeze_panes = "C2"
ws.auto_filter.ref = ws.dimensions

summary = wb.create_sheet("Summary", 0)
total = len(brands)
scraped_count = sum(1 for b in brands if scraped.get(b["slug"]) and not scraped[b["slug"]].get("delisted") and not scraped[b["slug"]].get("skipped") and not scraped[b["slug"]].get("error"))
delisted_count = sum(1 for b in brands if scraped.get(b["slug"], {}).get("delisted"))
skipped_count = sum(1 for b in brands if scraped.get(b["slug"], {}).get("skipped"))
stacking_flagged = sum(1 for b in brands if scraped.get(b["slug"]) and stacking_signal([
    clean(scraped[b["slug"]].get("Important Instructions", "")),
    clean(scraped[b["slug"]].get("Terms & Conditions", "")),
]) == "Yes")

summary.append(["Dealo / Checkout Assistant - Gyftr Brand Database"])
summary["A1"].font = Font(bold=True, size=14)
summary.append([])
for label, val in [
    ("Total brands", total),
    ("Fully scraped (T&C + Instructions + How to Use)", scraped_count),
    ("Delisted / dead pages on Gyftr", delisted_count),
    ("Skipped (utility/bill-pay, out of scope)", skipped_count),
    ("Brands flagged with possible stacking limits", stacking_flagged),
]:
    summary.append([label, val])
for row in summary.iter_rows(min_row=3, max_row=7):
    row[0].font = Font(bold=True)
summary.append([])
summary.append(["Note: 'Watch Video' links are pulled directly from Gyftr's own data and are not verified by us — some may be outdated or broken on their end."])
summary.append(["Note: 'Stacking Limit Mentioned?' is a keyword-based flag, not a parsed rule yet. 'Yes' means the T&C/Instructions text contains language suggesting a limit exists — always read the full text before assuming a specific number."])
summary.column_dimensions["A"].width = 45
summary.column_dimensions["B"].width = 12

wb.save("gyftr_full_database.xlsx")
print("Saved gyftr_full_database.xlsx")
print(f"Total: {total}, Scraped: {scraped_count}, Delisted: {delisted_count}, Skipped: {skipped_count}, Stacking-flagged: {stacking_flagged}")
